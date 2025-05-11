import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font

import utils

def build_workbook(tasks: list, team_db_file) -> str:
    wb = Workbook()
    wb.remove(wb.active)
    tasks_by_owner = utils.group_by_owner(tasks)
    for owner, items in tasks_by_owner.items():
        ws = wb.create_sheet(owner)
        df = pd.DataFrame(items)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        # Add styling here...
    output_path = "tasks_board.xlsx"
    wb.save(output_path)
    return output_path
